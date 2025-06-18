import xml.etree.ElementTree as ET
from xml.dom import minidom
from collections import OrderedDict
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

class ManagedObject:
    def __init__(self, mo_class, distname, version, mo_id, operation, paramdict):
        self.mo_class = mo_class
        self.distname = distname
        self.version = version
        self.mo_id = mo_id
        self.operation = operation
        self.paramdict = paramdict

class Document:
    def __init__(self, data_dict, xml_tree=None):
        self.data_dict = data_dict
        self.version_string = self.get_version_string()
        self.num_objects = len(data_dict)
        self.xml_tree = xml_tree

    def get_version_string(self):
        if self.data_dict:
            for key, params in self.data_dict.items():
                return params.get('_version', 'UNKNOWN')
        return 'UNKNOWN'

def make_name(st):
    l = st.split('/', 2)
    if len(l) != 3 or l[-1] in ('INTEGRATE-1','Device-1'):
        return False
    return l[-1]

def simplify_xml(root, ns):
    base = OrderedDict()
    for element in root.iterfind(".//ns:managedObject", ns):
        distName = element.attrib.get('distName')
        key = element.attrib.get('class')
        if not distName:
            continue
        if not key:
            continue
        base[key] = OrderedDict()
        base[key]['_class'] = element.attrib.get('class', 'UNKNOWN')
        base[key]['_distName'] = distName
        base[key]['_version'] = element.attrib.get('version', 'UNKNOWN')
        base[key]['_id'] = element.attrib.get('id', '10400')
        base[key]['_operation'] = element.attrib.get('operation', 'create')
        for parameters in element.findall("ns:p", ns):
            if parameters.text is not None:
                base[key][parameters.attrib['name']] = parameters.text
    return base

def build_full_xml(data_dict):
    NS_URI = "raml21.xsd"
    ET.register_namespace('', NS_URI)
    root = ET.Element("raml", {'version': '2.1', 'xmlns': NS_URI})
    cmData = ET.SubElement(root, "cmData", {
        'type': 'plan',
        'scope': 'all',
        'name': 'AIOSC-1-PnP-ProfileD-Basic-Integration-Planfile.xml'
    })

    common_version = get_version_string(data_dict)

    aiosc = ET.SubElement(cmData, "managedObject", {
        'class': "com.nokia.aiosc:AIOSC",
        'version': common_version,
        'distName': "PLMN-PLMN/AIOSC-6000039",
        'operation': "create"
    })
    ET.SubElement(aiosc, "p", {'name': "name"}).text = "PLMN-PLMN/AIOSC-6000039"
    ET.SubElement(aiosc, "p", {'name': "AutoConnHWID"}).text = "LBNKIASRC243920029"
    ET.SubElement(aiosc, "p", {'name': "$maintenanceRegionId"}).text = "PNP"
    ET.SubElement(aiosc, "p", {'name': "$maintenanceRegionCId"}).text = "1"
    ET.SubElement(aiosc, "p", {'name': "SparaPara2_CP"}).text = "1"
    ET.SubElement(aiosc, "p", {'name': "SparePara1_CP"}).text = "1"

    integrate = ET.SubElement(cmData, "managedObject", {
        'class': "com.nokia.integrate:INTEGRATE",
        'version': common_version,
        'distName': "PLMN-PLMN/AIOSC-6000039/INTEGRATE-1",
        'id': "104000",
        'operation': "create"
    })
    ET.SubElement(integrate, "p", {'name': "plannedSWReleaseVersion"}).text = "AIOSC24_01_400_35.aio.sig0"
    ET.SubElement(integrate, "p", {'name': "systemReleaseVersion"}).text = "AIOSC24"
    ET.SubElement(integrate, "p", {'name': "ipVersion"}).text = "0"

    device = ET.SubElement(cmData, "managedObject", {
        'class': "com.nokia.aiosc:Device",
        'version': common_version,
        'distName': "PLMN-PLMN/AIOSC-6000039",
        'operation': "create"
    })
    ET.SubElement(device, "p", {'name': "UserLabel"}).text = "AIOSC"

    for key, params in data_dict.items():
        current_params = params.copy()
        class_attr = current_params.pop('_class', 'UNKNOWN')
        dist_attr = current_params.pop('_distName', f"UNSPECIFIED/{key}")
        version = current_params.pop('_version', 'UNKNOWN')
        mo_id = current_params.pop('_id', '10400')
        operation = current_params.pop('_operation', 'create')
        mo = ET.SubElement(cmData, "managedObject", {
            'class': class_attr,
            'version': version,
            'distName': dist_attr,
            'id': mo_id,
            'operation': operation
        })
        for pname, val in current_params.items():
            ET.SubElement(mo, "p", {'name': pname}).text = val

    return ET.ElementTree(root)

def get_version_string(data_dict):
    if data_dict:
        first_key = next(iter(data_dict))
        return data_dict[first_key].get('_version', 'UNKNOWN')
    return 'UNKNOWN'

def make_xml(etree, docname):
    rough_string = ET.tostring(etree.getroot(), encoding="utf-8")
    pretty_xml = minidom.parseString(rough_string).toprettyxml(indent="    ")
    try:
        with open(f"{docname}.xml", "w", encoding="utf-8") as f:
            f.write(pretty_xml)
        print(f"XML file '{docname}.xml' generated successfully.")
    except IOError as e:
        print(f"Error writing XML file '{docname}.xml': {e}")


def update_dictionary(comp_base, comp_update, rename_dict):
    new_objs = []
    for key in list(comp_update.keys()):
        v = rename_dict.get(key, key)
        if v not in comp_base:
            mo = comp_update[key]
            new_objs.append([key, mo.get('_class', 'UNKNOWN'), mo.get('_distName', '')])
            continue

        mo_base = comp_base[v]
        mo_update = comp_update[key]
        for param in list(mo_update):
            if param.startswith('_'):
                continue
            param_rename = rename_dict.get(param, param)
            if param_rename in mo_base:
                mo_update[param] = mo_base[param_rename]
    return comp_update, new_objs

def generate_excel_report(comp_base, comp_update, new_objs, report_filename="AIOSC_comparison_report.xlsx"):
    """
    Generates a single Excel (.xlsx) file with a formatted table and summary using openpyxl.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison Report"

    all_objects_for_report = []

    new_obj_keys = {item[0] for item in new_objs}

    # Populate all_objects_for_report with status and relevant data
    # Start with new objects
    for key, mo_data in comp_update.items():
        if key in new_obj_keys:
            all_objects_for_report.append({
                "Status": "NEW",
                "Class": mo_data.get('_class', 'UNKNOWN'),
                "DistName": mo_data.get('_distName', '')
            })

    # Add deprecated objects
    deprecated_keys = []
    for key in comp_base:
        if key not in comp_update:
            deprecated_keys.append(key)
            mo_data = comp_base[key]
            all_objects_for_report.append({
                "Status": "DEPRECATED",
                "Class": mo_data.get('_class', 'UNKNOWN'),
                "DistName": mo_data.get('_distName', '')
            })

    # Add common objects
    for key, mo_data in comp_update.items():
        if key in comp_base and key not in new_obj_keys and key not in deprecated_keys:
             all_objects_for_report.append({
                "Status": "COMMON",
                "Class": mo_data.get('_class', 'UNKNOWN'),
                "DistName": mo_data.get('_distName', '')
            })

    # Define headers for the data table
    data_headers = ["Status", "Class", "DistName"]
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center')

    # --- Summary Section ---
    current_row = 1
    ws.cell(row=current_row, column=1, value="--- Comparison Summary ---").font = header_font
    current_row += 2 # Skip a row for spacing

    ws.cell(row=current_row, column=1, value="Metric").font = header_font
    ws.cell(row=current_row, column=2, value="Value").font = header_font
    current_row += 1

    ws.cell(row=current_row, column=1, value="Total Objects in Base Version")
    ws.cell(row=current_row, column=2, value=len(comp_base))
    current_row += 1

    ws.cell(row=current_row, column=1, value="Total Objects in Update Version")
    ws.cell(row=current_row, column=2, value=len(comp_update))
    current_row += 1

    ws.cell(row=current_row, column=1, value="New Objects Added")
    ws.cell(row=current_row, column=2, value=len(new_objs))
    current_row += 1

    ws.cell(row=current_row, column=1, value="Objects Deprecated")
    ws.cell(row=current_row, column=2, value=len(deprecated_keys))
    current_row += 2 # Skip a row for spacing

    # --- Main Data Table Heading ---
    ws.cell(row=current_row, column=1, value="--- Detailed Object Comparison ---").font = header_font
    current_row += 2 # Skip a row for spacing

    # Write data headers
    for col_idx, header in enumerate(data_headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = center_align
    current_row += 1

    # Sort data for better grouping
    sorted_objects = sorted(
        all_objects_for_report,
        key=lambda x: (
            0 if x["Status"] == "NEW" else
            1 if x["Status"] == "DEPRECATED" else
            2
        )
    )

    # Write sorted data
    for row_data in sorted_objects:
        ws.cell(row=current_row, column=1, value=row_data["Status"])
        ws.cell(row=current_row, column=2, value=row_data["Class"])
        ws.cell(row=current_row, column=3, value=row_data["DistName"])
        current_row += 1

    # Auto-adjust column widths for readability
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) # Add a little padding
        ws.column_dimensions[column].width = adjusted_width

    try:
        wb.save(report_filename)
        print(f"Unified comparison report written to '{report_filename}'")
    except Exception as e:
        print(f"Error writing Excel report: {e}")

if __name__ == "__main__":
    rename_dict = {}
    ns = {'ns': 'raml21.xsd'}

    xml_file_update = "AIOSC25_drop1_dataModel.xml"
    xml_file_base = "Nokia_AIOSC24_SCF_NIDD4.0_v17.xml"

    print(f"Attempting to load XML files:")
    print(f"  Base File: '{xml_file_base}'")
    print(f"  Update File: '{xml_file_update}'")

    try:
        if not os.path.exists(xml_file_update):
            raise FileNotFoundError(f"'{xml_file_update}' does not exist.")
        if not os.path.exists(xml_file_base):
            raise FileNotFoundError(f"'{xml_file_base}' does not exist.")

        tree_update = ET.parse(xml_file_update).getroot()
        tree_base = ET.parse(xml_file_base).getroot()
    except FileNotFoundError as e:
        print(f"Error: {e} Please ensure the XML files are in the same directory as the script.")
        exit(1)
    except ET.ParseError as e:
        print(f"Error parsing XML file: {e}. Please check if the XML files are well-formed.")
        exit(1)

    comp_base = simplify_xml(tree_base, ns)
    comp_update = simplify_xml(tree_update, ns)
    print(comp_base)

    update_dict, new_objs = update_dictionary(comp_base, comp_update, rename_dict)
    comp_finalfile = build_full_xml(update_dict)

    docname = input("Please enter the desired file name for the generated XML (e.g., 'MyOutputPlan'): ")
    make_xml(comp_finalfile, docname)

    print("\n--- Console Comparison Summary ---")
    print(f"Base Version Objects: {len(comp_base)}")
    print(f"Update Version Objects: {len(comp_update)}")

    num_new_objects = len(new_objs)
    num_deprecated_objects = len([key for key in comp_base if key not in comp_update])

    print(f"Number of **NEW** objects introduced: {num_new_objects}")
    print(f"Number of **DEPRECATED** objects (removed from base): {num_deprecated_objects}")
    print("----------------------------------")

    # Generate a single, unified Excel report
    generate_excel_report(comp_base, comp_update, new_objs)

    base_document = Document(comp_base, tree_base)
    update_document = Document(comp_update, comp_finalfile)